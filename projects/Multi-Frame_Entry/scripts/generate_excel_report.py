#!/usr/bin/env python3
"""
生成多品种对比Excel报告

功能：
1. 读取所有品种的训练结果
2. 生成多sheet Excel文件
3. 每个品种一个详细结果sheet
4. 最后一个Summary sheet包含跨品种对比
"""
import pandas as pd
from pathlib import Path
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

# 样式配置
HEADER_STYLE = {
    'font': Font(bold=True, size=12, color='FFFFFF'),
    'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
    'alignment': Alignment(horizontal='center', vertical='center'),
    'border': Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
}

SUBHEADER_STYLE = {
    'font': Font(bold=True, size=11),
    'fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
    'alignment': Alignment(horizontal='center', vertical='center')
}

DATA_STYLE = {
    'alignment': Alignment(horizontal='center', vertical='center'),
    'border': Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
}


def generate_excel_report(
    results_dir: Path,
    output_file: Path = None
):
    """
    生成Excel报告

    Args:
        results_dir: 训练结果目录
        output_file: 输出文件路径
    """
    logger.info("\n生成Excel报告...")

    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = results_dir / f'multi_symbol_comparison_{timestamp}.xlsx'

    # 读取训练汇总
    summary_file = results_dir / 'training_summary.csv'
    if not summary_file.exists():
        logger.error(f"训练汇总文件不存在: {summary_file}")
        return None

    df_summary = pd.read_csv(summary_file)

    # 创建Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # ==================== Summary Sheet ====================
        df_summary_pivot = df_summary.pivot(
            index='品种',
            columns='年份',
            values='AUC'
        )

        # 添加平均列
        df_summary_pivot['平均AUC'] = df_summary_pivot.mean(axis=1)

        # 保存Summary
        df_summary_pivot.to_excel(writer, sheet_name='Summary')

        # ==================== 各品种详细Sheet ====================
        for symbol in df_summary['品种'].unique():
            df_symbol = df_summary[df_summary['品种'] == symbol]

            # 创建详细信息表
            detail_data = []
            for _, row in df_symbol.iterrows():
                year = int(row['年份'])
                detail_data.append({
                    '指标': f'{year}年',
                    '子指标': '准确率',
                    '值': float(row['准确率'])
                })
                detail_data.append({
                    '指标': f'{year}年',
                    '子指标': 'AUC',
                    '值': float(row['AUC'])
                })
                detail_data.append({
                    '指标': f'{year}年',
                    '子指标': 'F1分数',
                    '值': float(row['F1'])
                })

            df_detail = pd.DataFrame(detail_data)
            df_detail.to_excel(writer, sheet_name=symbol.replace('.', '_'), index=False)

    # ==================== 格式化 ====================
    logger.info("  格式化Excel文件...")
    wb = load_workbook(output_file)

    # 格式化Summary sheet
    ws_summary = wb['Summary']
    _format_sheet(ws_summary, df_summary_pivot)

    # 格式化各品种sheet
    for symbol in df_summary['品种'].unique():
        sheet_name = symbol.replace('.', '_')
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _format_symbol_sheet(ws, symbol, df_symbol)

    # 添加标题行到Summary
    ws_summary.insert_rows(1)
    ws_summary['A1'] = '多品种滚动训练对比 - AUC汇总'
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    ws_summary.merge_cells('A1:G1')

    wb.save(output_file)
    logger.info(f"  ✓ Excel报告已保存: {output_file}")

    return output_file


def _format_sheet(ws, df):
    """格式化Summary sheet"""
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    for col in range(2, len(df.columns) + 2):
        ws.column_dimensions[chr(64 + col)].width = 12

    # 格式化表头
    for cell in ws[2]:
        cell.font = HEADER_STYLE['font']
        cell.fill = HEADER_STYLE['fill']
        cell.alignment = HEADER_STYLE['alignment']
        cell.border = HEADER_STYLE['border']

    # 格式化数据
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = DATA_STYLE['alignment']
            cell.border = DATA_STYLE['border']
            # 高AUC值标绿
            if isinstance(cell.value, float) and cell.value > 0.6:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            # 低AUC值标红
            elif isinstance(cell.value, float) and cell.value < 0.5:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')


def _format_symbol_sheet(ws, symbol, df_symbol):
    """格式化品种sheet"""
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    # 添加标题
    ws.insert_rows(1)
    ws['A1'] = f'{symbol} 详细性能指标'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:C1')

    # 格式化表头
    for cell in ws[3]:
        cell.font = HEADER_STYLE['font']
        cell.fill = HEADER_STYLE['fill']
        cell.alignment = HEADER_STYLE['alignment']
        cell.border = HEADER_STYLE['border']

    # 格式化数据
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = DATA_STYLE['alignment']
            cell.border = DATA_STYLE['border']


def main():
    """主函数"""
    # 路径配置（基于脚本位置）
    project_root = Path(__file__).parent.parent
    results_dir = project_root / 'models' / 'training_results'
    output_file = results_dir / 'multi_symbol_comparison.xlsx'

    # 验证结果目录
    if not results_dir.exists():
        print(f"\n✗ 训练结果目录不存在: {results_dir}")
        print("   请先运行训练脚本")
        return None

    excel_file = generate_excel_report(results_dir, output_file)

    if excel_file:
        print(f"\n✓ Excel报告生成成功: {excel_file}")
    else:
        print("\n✗ Excel报告生成失败")

    return excel_file


if __name__ == '__main__':
    main()
